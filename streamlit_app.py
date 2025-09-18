# -*- coding: utf-8 -*-
import streamlit as st, pandas as pd
from io import BytesIO
from transformer import transform

st.set_page_config(page_title="ReportCaster — Form → Fonksiyonlar", layout="wide")
st.title("📑 ReportCaster — Form → 📊 Fonksiyonlar Data")
st.caption("Birebir şablon denemesi + programatik Dashboard (Excel içinde).")

with st.sidebar:
    st.header("Ayarlar")
    faz_value = st.text_input("Faz", value="Faz 6")
    st.write("İsteğe bağlı: Şablon Excel (2.xlsx) yükleyin, aksi halde sadece programatik dashboard üretilir.")
    tpl_file = st.file_uploader("Şablon (2.xlsx)", type=["xlsx"], key="tpl")

src_file = st.file_uploader("1 numaralı formatta Excel", type=["xlsx"], key="src")

if not src_file:
    st.info("Başlamak için kaynak Excel yükleyin.")
    st.stop()

xls = pd.ExcelFile(src_file)
df_raw = pd.read_excel(xls, sheet_name=xls.sheet_names[0])
st.subheader("Kaynak örnek")
st.dataframe(df_raw.head(20), use_container_width=True)

out_df = transform(df_raw, faz_value=faz_value)
st.subheader("Dönüştürülmüş veri (önizleme)")
st.dataframe(out_df.head(50), use_container_width=True)

# 1) Programatik dashboard (xlsxwriter) — ŞABLONSUZ, her ortamda çalışır
st.markdown("### 📦 Dashboard'lu Excel (şablonsuz)")
import numpy as np
buffer_dash = BytesIO()
with pd.ExcelWriter(buffer_dash, engine="xlsxwriter") as writer:
    out_df.to_excel(writer, index=False, sheet_name="Fonksiyonlar Data")
    df = out_df.copy()
    df["Tarih"] = pd.to_datetime(df["Tarih"], errors="coerce")
    df["Yıl-Ay"] = df["Tarih"].dt.to_period("M").astype(str)

    agg1 = df.groupby(["Yıl-Ay","Uygulama"], dropna=False)["Puan"].mean().reset_index()
    agg2 = df.groupby(["Uygulama"], dropna=False)["Puan"].mean().reset_index()
    agg3 = df.groupby(["Test Alanı"], dropna=False)["Puan"].mean().sort_values("Puan", ascending=False).reset_index()

    agg1.to_excel(writer, sheet_name="DashboardData", index=False, startrow=0, startcol=0)
    agg2.to_excel(writer, sheet_name="DashboardData", index=False, startrow=0, startcol=5)
    agg3.to_excel(writer, sheet_name="DashboardData", index=False, startrow=0, startcol=9)

    wb = writer.book
    ws_dash = wb.add_worksheet("Dashboard")
    title = wb.add_format({"bold": True, "font_size": 16})
    ws_dash.write(0, 0, "Rapor Özeti", title)

    chart1 = wb.add_chart({"type": "line"})
    n1 = len(agg1)
    if n1 > 0:
        chart1.add_series({
            "name": "Aylık Ortalama",
            "categories": f"=DashboardData!$A$2:$A${n1+1}",
            "values":     f"=DashboardData!$C$2:$C${n1+1}",
        })
        chart1.set_title({"name": "Aylık Ortalama Puan"})
        ws_dash.insert_chart(2, 0, chart1, {"x_scale": 1.2, "y_scale": 1.2})

    chart2 = wb.add_chart({"type": "column"})
    n2 = len(agg2)
    if n2 > 0:
        chart2.add_series({
            "name": "Uygulama Ort.",
            "categories": f"=DashboardData!$F$2:$F${n2+1}",
            "values":     f"=DashboardData!$G$2:$G${n2+1}",
        })
        chart2.set_title({"name": "Uygulama Bazlı Ortalama Puan"})
        ws_dash.insert_chart(20, 0, chart2, {"x_scale": 1.2, "y_scale": 1.2})

    chart3 = wb.add_chart({"type": "bar"})
    n3 = len(agg3)
    if n3 > 0:
        chart3.add_series({
            "name": "Test Alanı Ort.",
            "categories": f"=DashboardData!$J$2:$J${n3+1}",
            "values":     f"=DashboardData!$K$2:$K${n3+1}",
        })
        chart3.set_title({"name": "Test Alanı Bazlı Ortalama Puan"})
        ws_dash.insert_chart(2, 8, chart3, {"x_scale": 1.2, "y_scale": 1.2})

buffer_dash.seek(0)
st.download_button("⬇️ Dashboard'lu Excel (indir)", data=buffer_dash, file_name="rapor_dashboard.xlsx")

# 2) Birebir şablona yaz (deneysel) — pivot/slicer’lı şablonlarda bazen sorun olabilir
st.markdown("### 🧩 Birebir Şablon Yaz (deneysel)")
if tpl_file:
    try:
        tpl_bytes = tpl_file.read()
        from openpyxl import load_workbook
        wb = load_workbook(BytesIO(tpl_bytes), keep_links=False)  # pivot/slicer hatalarını çoğunlukla by-pass eder
        wsname = "Fonksiyonlar Data"
        if wsname not in wb.sheetnames:
            st.error("Şablonda 'Fonksiyonlar Data' sayfası yok.")
        else:
            ws = wb[wsname]
            if ws.max_row > 1:
                ws.delete_rows(2, ws.max_row-1)
            template_headers = [cell.value for cell in ws[1]]
            cols = template_headers if all(template_headers) else out_df.columns.tolist()
            for c in cols:
                if c not in out_df.columns:
                    out_df[c] = None
            out = out_df[cols]
            for r_idx, (_, r) in enumerate(out.iterrows(), start=2):
                for c_idx, col in enumerate(cols, start=1):
                    ws.cell(row=r_idx, column=c_idx, value=r.get(col, None))
            out_bytes = BytesIO()
            wb.save(out_bytes)
            out_bytes.seek(0)
            st.download_button("⬇️ Şablona Yazılmış Excel (indir)", data=out_bytes, file_name="rapor_birebir.xlsx")
    except Exception as e:
        st.warning(f"Şablon açılamadı: {e}")
        st.info("Üstteki 'Dashboard'lu Excel her koşulda çalışır.")
else:
    st.info("Birebir şablon için şablon dosyanızı yükleyin.")
