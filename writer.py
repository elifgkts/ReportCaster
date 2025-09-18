# -*- coding: utf-8 -*-
from __future__ import annotations
from io import BytesIO
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

def write_into_template(template_bytes: bytes,
                        fonk_df: pd.DataFrame,
                        up_df: pd.DataFrame,
                        fonk_sheet_name: str = "Fonksiyonlar Data",
                        up_sheet_name: str = "UploadDownload Data") -> bytes:
    """
    Şablon workbook'u açar:
      - İki sayfadaki (başlık satırı sabit) veri alanlarını temizler,
      - Yeni veriyi yazar,
      - Sayfadaki Excel Table'ların ref aralıklarını otomatik büyütür.
    Biçimler, filtreler, pivot/slicer vb. şablondaki gibi kalır.
    """
    wb = load_workbook(BytesIO(template_bytes), keep_links=False, data_only=False)

    # ----- Fonksiyonlar Data -----
    if fonk_sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{fonk_sheet_name}' not found in template.")
    ws = wb[fonk_sheet_name]
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row-1)
    template_headers = [c.value for c in ws[1]]
    cols = template_headers if all(template_headers) else list(fonk_df.columns)
    for c in cols:
        if c not in fonk_df.columns: fonk_df[c] = None
    fonk_out = fonk_df[cols]
    for r_idx, (_, r) in enumerate(fonk_out.iterrows(), start=2):
        for c_idx, col in enumerate(cols, start=1):
            ws.cell(row=r_idx, column=c_idx, value=r.get(col))
    if ws._tables:
        for t in ws._tables.values():
            last_row = max(2, fonk_out.shape[0] + 1)
            last_col = fonk_out.shape[1]
            t.ref = f"A1:{get_column_letter(last_col)}{last_row}"

    # ----- UploadDownload Data -----
    if up_sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{up_sheet_name}' not found in template.")
    ws2 = wb[up_sheet_name]
    if ws2.max_row > 1:
        ws2.delete_rows(2, ws2.max_row-1)
    template_headers2 = [c.value for c in ws2[1]]
    cols2 = template_headers2 if all(template_headers2) else list(up_df.columns)
    for c in cols2:
        if c not in up_df.columns: up_df[c] = None
    up_out = up_df[cols2]
    for r_idx, (_, r) in enumerate(up_out.iterrows(), start=2):
        for c_idx, col in enumerate(cols2, start=1):
            ws2.cell(row=r_idx, column=c_idx, value=r.get(col))
    if ws2._tables:
        for t in ws2._tables.values():
            last_row = max(2, up_out.shape[0] + 1)
            last_col = up_out.shape[1]
            t.ref = f"A1:{get_column_letter(last_col)}{last_row}"

    bio = BytesIO()
    wb.save(bio); bio.seek(0)
    return bio.read()

def write_portable_with_tables(fonk_df: pd.DataFrame,
                               up_df: pd.DataFrame,
                               fonk_sheet_name: str = "Fonksiyonlar Data",
                               up_sheet_name: str = "UploadDownload Data") -> bytes:
    """
    xlsxwriter ile iki sayfalı yeni bir dosya oluşturur ve
    her sayfada Excel Table (filtreli) kurar.
    """
    from io import BytesIO
    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine="xlsxwriter") as writer:
        fonk_df.to_excel(writer, index=False, sheet_name=fonk_sheet_name)
        up_df.to_excel(writer, index=False, sheet_name=up_sheet_name)
        wb = writer.book
        ws = writer.sheets[fonk_sheet_name]
        ws2 = writer.sheets[up_sheet_name]
        nrows, ncols = fonk_df.shape
        ws.add_table(0, 0, max(1, nrows), max(0, ncols-1), {
            "name": "FonksiyonlarData",
            "columns": [{"header": c} for c in fonk_df.columns],
            "autofilter": True
        })
        n2, m2 = up_df.shape
        ws2.add_table(0, 0, max(1, n2), max(0, m2-1), {
            "name": "UploadDownloadData",
            "columns": [{"header": c} for c in up_df.columns],
            "autofilter": True
        })
    buffer.seek(0)
    return buffer.read()
